#count word-based features in URL 

import openpyxl
from openpyxl import Workbook
import nltk
from nltk.tokenize import sent_tokenize, wordpunct_tokenize

# Give the location of the file 
# path = "gfg.xlsx"
  
# To open the workbook 
# workbook object is created 
wb_obj = openpyxl.load_workbook('../../Data_Collection_03/02_Preprocessing/02_URL/04_url_all.xlsx') 
  
# Get workbook active sheet object 
# from the active attribute 
sheet_obj = wb_obj.active 

row = sheet_obj.max_row
column = sheet_obj.max_column
rowcell = "A"+str(row)

# Cell object is created by using 
# sheet object's cell() method. 
cell_obj = sheet_obj['A1': rowcell]

blog = ['blogspot','blog','wordpress','blogger']
socmed = ['facebook', 'twitter', 'instagram']
news = ['news','press','newspress','journal','publisher']
covid = ['coronavirus','virus','covid']

i = 1

for cell1 in cell_obj:

    
    c1 = sheet_obj.cell(row = i, column = 1)
    c2 = sheet_obj.cell(row = i, column = 2)
    c3 = sheet_obj.cell(row = i, column = 3)
    c4 = sheet_obj.cell(row = i, column = 4)
    c5 = sheet_obj.cell(row = i, column = 5)
    words = wordpunct_tokenize(c1.value)
    
    for w in words:

        if w in blog:
            c2.value = True
        else:
            if not c2.value == True:
                c2.value = False
      
        if w in socmed:
            c3.value = True
        else:
            if not c3.value == True:
                c3.value = False
            
        if w in news:
            c4.value = True
        else:
            if not c4.value == True:
                c4.value = False
            
        if w in covid:
            c5.value = True
        else:
            if not c5.value == True:
                c5.value = False
    
    i=i+1


wb_obj.save("../../Data_Collection_03/02_Preprocessing/02_URL/05_url_wordbased_trash.xlsx")



    