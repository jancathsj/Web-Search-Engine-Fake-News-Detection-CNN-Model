#remove stopwords and special characters
#count double quotation marks
  
import openpyxl 
#import re
import nltk
from nltk.tokenize import sent_tokenize, word_tokenize
from nltk.corpus import stopwords
#nltk.download('stopwords')
#nltk.download('punkt')
  
# Give the location of the file 
# path = "gfg.xlsx"
  
# To open the workbook 
# workbook object is created 
wb_obj = openpyxl.load_workbook('../../Data_Collection_03/02_Preprocessing/01_TitleDesc/01_TitleDesc.xlsx')
#wb_obj = openpyxl.load_workbook('Dataset_words.xlsx') 
#wb_obj = openpyxl.load_workbook('Sample_words.xlsx') 
  
# Get workbook active sheet object 
# from the active attribute 
sheet_obj = wb_obj.active 

row = sheet_obj.max_row
column = sheet_obj.max_column
rowcell = "B"+str(row)

# Cell object is created by using 
# sheet object's cell() method. 
cell_obj = sheet_obj['A1': rowcell]

wordcount = 0
qcountt = 0
qcountd = 0
stopWords = nltk.corpus.stopwords.words('english')
newstopWords=['accordance', 'according', 'also', 'another','claim','comprises','corresponding','could','described','desired','embodiment','fig','figs','generally','herein','however','invention','means','onto','particularly','preferably','preferred','present','provide','provided','provides','relatively','respectively','said','since','suitable','then','thereby','therefore','thereof','thereto','thus','use','various','whereby','wherein','would','able','mentioned','accordingly','across','along','already','alternatively','among','anywhere','better','disclosure','due','easily','easy','eg','either','elsewhere','enough','especially','essentially','et al','et','al','etc','eventually','excellent','finally','furthermore','good','hence','ie','ii','iii','instead','later','like','little','many','may','meanwhile','might','moreover','much','must','never','often','others','otherwise','overall','rather','remarkably','significantly','simply','sometimes','specifically','straight','forward','substantially','thereafter','therebetween','therefor','therefrom','therein','thereinto','thereon','therethrough','therewith','together','toward','towards','typical','upon','via','vice versa','whatever','whereas','whereat','wherever','whether','whose','within','without','yet']
stopWords.extend(newstopWords)

#foreign =  [ 'rybicki', 'narvaez', 'nassau', 'curran', 'loppi', 'soave', 'ke', 'sachiha', 'kurose', 'nikko', 'masaki', 'hirano', 'masaya', 'kato', 'taisei', 'hoyama', 'kiran', 'sharma', 'shunsuke', 'tabeta', 'cascella', 'auntyuta', 'sinan', 'i̇şkur', 'emeka', 'oparah', 'padma', 'shri', 'paese', 'inuvik', 'kami', 'yash', 'jong', 'tobago', 'gaddafi', 'elibertadores', 'kombat', 'buhari', 'melaye', 'amharic', 'ebtedge', 'grbic', 'bih', 'nadia', 'bouzid', 'kamome', 'chuo', 'marino', 'jal', 'kono', 'exgp', 'mdka', 'vkdgqu', 'pjqvqu', 'exgp201', 'pna', 'gtu', 'plazest', 'tj', 'hqnnqykpi', 'rtqokpgp', 'kpdkxkdwanu', 'htqo', 'inban', 'nomi', 'esri', 'pr', 'italyjapanmiddle', 'eastmyanmarnepalnew', 'zealandpakistanrussiasingaporesouth', 'doori', 'kalyan-dombivli', 'japanmiddle', 'netripura', 'akrspi', 'manohar', 'parrikar', 'shahid-beheshti', 'kuchikomi', 'higashi-kanda', 'chiyoda-ku', 'shinkansen', 'ryokan', 'onsen', 'shoko', 'fujimisou', 'nishiura', 'aichi', 'filed', 'nagoya', 'atsumi', 'abs', 'dmss', 'moi', 'ro', 'kiraannjapan', 'lao', 'liosatos', 'aflw', 'tayla', 'suvamjyoti', 'ffrs', 'mhlw', 'amabie', 'hoshino', 'yomiuriland', 'fuji-q', 'kazuhiro', 'ota', 'mainichi/kota', 'yoshida', 'jhu', "'sabungeros", 'fani-kayode', 'ffk', 'que', 'sisay', 'woubeshet', 'lafave-ransom', 'ont', 'hobkcirc', 'kokomo', 'iupuc', 'iupui', 'saginaw', 'siliguri', 'kolkata', 'hiroshi', 'tabata', 'yuriko', 'koike', 'cuomo', 'kerala', 'pak', 'lanka', 'ビル・ゲイツ＝イベント', 'コロナウイルス特許', 'manak', 'motomami', 'rosalía', 'tewari', 'flugge', 'tecumseh', 'gamberini', 'amta', 'anadolu', 'ajansı','dziedzic', 'morawska', 'elevit', '新冠假病毒', 'bayanihan', 'cereda', 'yuli', 'roma', 'ahmetovic', 'portese', 'sardines', 'denis', 'mukwege', 'lise', 'aserud/epa', 'ikaba', 'koyi', 'getafe', 'yamato', 'amhara', 'kii', 'moreno', 'fceu001', 'sunak', 'niro',  'suffolk', 'magnifica', 'sivrice', 'elazig', 'maio', 'sveriges', 'riksbank', 'kiyoshi', 'renationalize', 'alitalia', 'smeralda', 'seckin/anadolu', 'narendra',  'maharashtra', 'ahla', 'ryozo', 'matsuda', 'daszak', 'rosato', 'palisades', 'bourouiba', 'rebbe', 'menachem', 'schneerson', 'obm', 'rei', 'kokotel', 'shihab', 'abates', 'argonne', 'bergoglio', 'pachamama', 'brezina', 'stitt', 'acces', 'vrtojba','eroshenko', 'rajaniemi', 'huileng', 'karnataka', 'navami', 'tata', 'srivastava', 'ayon', 'mukhopadhyay', 'iifl', 'varun', 'lohchab', 'aditya-l1', 'amon', 'santwon', "'maskne", 'nimmo', 'kilmarnock', 'lanarkshire', 'deccan', 'mayank', '472-2189', '©', 'moura', 'schilingi', 'roness', 'yau', 'neuraminidase', 'setti', 'borje', 'ekholm', 'tmims', 'mahmoud', 'ahmadinejad', '🤩', 'aml/cft', 'fbc', 'fcc', 'ajit', 'pai', 'lffll', 'igg', '06r90', 'h14814r06', 'impella', 'hemashield', '3.7', 'kinji', 'fukasaku', 'chandigarh', 'akhilesh', 'yadav', 'lucknow', 'samajwadi', 'lashed', 'uttar', 'pradesh', 'pellettiere', 'marazzi', 'hinodecho', 'pareto', 'malpensa', 'alphatauri', 'eau', 'didier', 'raoult','larouche', 'sonesta','bozen', 'bolzano', 'adige', 'südtirol', 'hek-293t', 'cui', 'aagaa', 'batsona', 'tema', 'sabungeros', 'ng', 'estudyante', 'natunton', 'nang', 'kuhaan', 'nagmalasakit', 'motorista', 'iag', 'antim', 'panghal', 'anubha', 'rohatgi', 'aptima', 'srinagar', 'oromia', 'depo-provera', 'navrongo', 'daegu', 'odisha', 'kochi', 'rajnath', 'djokovic', 'jelena', 'terre', 'inss', 'jerram', 'rvcjinsta', 'ravi', 'teja', 'orbán', "orbán's",  'shabak', 'yitzhak', 'ilan', 'zweli', 'mkhize', 'yubari', 'yasutoshi', 'nishimura', 'kazuyoshi', 'akaba', 'ziona', 'ljungblad', 'nyumbakumi', 'onyancha', 'buddhipongse', 'jaisalmer', 'madhya', 'ayatollah', 'hashem', 'bathayi', 'golpayegani', 'shoham', 'raakesh', 'saraff', 'hofstede', 'masayoshi', 'ritesh', 'bipin', 'rawat', 'manoj', 'mukund', 'naravane', 'kahan', 'cellecor', 'holmarc', 'kassam', 'abela', 'tzu', 'sangsahachart/', '経済産業省', 'sterrad', '健康・医療covid-19', 'nora', 'gámez', 'nuevo', 'mena', 'malfatto', 'mondaq', 'arpita.padiyar', 'buziashvili', 'dasgupta', 'nami', 'mongkut', 'ladkrabang', 'kmitl', 'qui', 'clodomir', 'santana', 'federico', 'botta', 'riccardo', 'clemente', 'licoppe', 'bolsonaro', 'merck', 'niels', 'finsen', 'bei', 'ashish', 'kurme', 'mbbs', 'drescher', 'sorkin', 'rossi', 'amihilda', 'menina', 'normina', 'nicotra', 'mendoza', 'evstatieva', 'fakta', 'kenapa', 'rakyat', 'memandu', 'sebelah', 'berhad', 'luci', 'hiroko', 'masuike/the', 'novo-ogaryovo', 'burgum', 'chien-jen', 'arevalo', 'jin', 'weiyi', 'vaporetto', 'poehler', 'ep0369695a2', 'ca2187704a1', 'ca3102135a1', 'ca3129725a1', 'ca2460367a1', 'ep3931771a1', 'cn105917404b', 'us7089208b1', 'us6226615b1', 'us9498694b2', 'sabaudia', 'tiyan', 'ospital', 'pajanel', 'subacchi', 'cavalieri', 'leclerc', 'minh', 'mach', 'phuong', 'khanh', 'niraj', 'pandit', 'webasto', 'jörn', 'poltz', 'riham', 'alkousaa', 'lng', 'joox', 'samar', 'hassan', 'rania', 'gamal', 'tmsnrt.rs/2rbwi5e', 'kabul', 'ghani', 'escondida', 'pampa', 'norte', 'scholz', 'dayaram', 'pirhossein', 'kolivand', 'ilna', 'misiani', 'zingaretti', 'kosovo', 'ringgit', 'rebelo', 'sousa', 'ramaphosa', 'hsien', 'ligue', 'mpango', 'phuc', 'bahn', 'tasuku', 'honjo', 'sanofi', "d'italia", 'matignon', '\u202e201\u202c', 'reuters/tatyana', 'makeyeva', 'stina', 'anshuman', 'daga', 'himani', 'navaratnam', 'alli', 'mallorca', 'addis', 'ababa', 'jaafar', 'allawi', 'azar',  'barbuscia', 'saba', 'duran', 'nyse', 'harare', 'alibaba', 'jomo', 'kenyatta', 'tirana', 'leke', 'anila', 'denaj', 'abdelaziz', 'djerad', 'buenos', 'aires', 'yerevan', 'canberra', 'dutton', 'kurz', 'meeus', 'sarajevo', 'patrasso/file', 'barata', 'paraisopolis', 'favelas', 'militias', 'janeiro', 'reuters/ian', 'cheibub', 'paulo/geneva', 'phnom', 'reuters/tingshu', 'suifenhe', 'heilongjiang', 'reuters/huizhong', 'bogota', 'broendby', 'ritzau', 'scanpix/martin', 'arcos', 'quito', 'reuters/tiksa', 'negeri', 'rutte', 'sureema', 'suarez', 'exxon', 'lehtikuva/heikki', 'saukkomaa', 'gakharia', 'reuters/lisi', 'niesner', 'babis', 'kreuzberg', 'hanschke', 'staikouras', 'maximos', 'papamitsos/handout', 'tegucigalpa', 'cabrera', 'schmuelgen', 'bhiwandi', 'mascarenhas', 'kurniawan', 'reuters/ajeng', 'ulfiana', 'tangerang', 'sukoharjo', 'rouhani', 'reuters/remo', 'mafiosi', 'reuters/issei', 'tanaka', 'fujitsu', 'leussink', 'makiko', 'tokayev', 'uhuru', 'nassif', 'hitti', 'reuters/darrin', 'zammit', 'lupi', 'reuters/lucas', 'harlingen', 'novacyt', 'reuters/stoyan', 'nenov', 'lima', 'vizcarra', 'qatari', 'nascimento', 'quilombo', 'reuters/evgenia', 'novozhenina', 'chechnya', 'kigali', 'bizimana', 'rwandan', 'rya.i', 'qatif', 'yosri', 'shionogi', 'reuters/dado', 'rapinoe','nwsl', 'reuters/sivaram', 'reuters/rupak', 'grinon', 'moncloa', 'cadenas',  'reuters/dinuka', 'liyanawatte/file', 'kimali', 'tantolunden','thais', 'reuters/athit', 'perawongmetha', 'reuters/umit', 'bektas/file', 'erdogan', 'chp', 'tayyip', 'yoweri', 'museveni', 'pelosi', 'ola', 'dji', 'hivju', 'zocalo', 'pereta', 'aaas', 'kentaro', 'iwata', 'ssleec', 'widodo', 'zuma', 'bonnier', 'duomo', 'saey', 'pasar', 'pramuka',  'aafes', 'garganera', 'onondaga', 'boko', 'fanos', 'panayides', 'nazia', 'parveen', 'jusna', 'ilda', 'harikumar','dagenham', 'anggono/sas/20', 'pindad', 'nisar', 'assoturismo', 'confesercenti', 'vittorio', 'messina', 'schwartzapfel', 'alexievich', 'svetlana', 'pathanamthitta',  'tacoma', 'inslee', 'jharkhand', 'nwo', 'warubuko', 'sokoke_khadzonzo', 'marat', 'safin', 'carletta', 'puglia', 'oia', 'izakaya', 'shinbashi','الأمم', 'المتحدة', 'maas', 'havfarm', 'unfpa', 'palais', 'raparin', 'belinelli', 'oann', 'biolo', 'mauritius', 'corriere', 'della', 'pattugalan', 'maj.', 'pengbai', 'obaseki', 'tamm', 'geraint', 'teikoku', 'asif', 'auken', 'leazer', 'wwny', 'gfl', 'cesare', 'jabil', 'zócalo', 'nespolon', 'kujawski', 'alina', 'bârgăoanu','abuja', 'lagos', 'beto', 'rongen', 'baliji', 'zust', 'vardhan', 'limbaugh', 'manzanar','ahmedabad', 'ahmedbad', 'fouchier', 'erasmus', 'zulfiqar', 'nader', 'ibrahim', 'lazara', 'marinkovic', 'mwai', 'olga', 'shayan', 'sardarizadeh', 'bregadze', 'lugar',  'giraldi', 'koblentz', 'kortepeter', 'cefalu',  'nikai', 'wakayama', 'lakh', 'ibaraki', 'rashid', 'mahesh', 'rebelde', 'estévez', 'aiirs', 'pansini', 'fornacca', 'conticini', 'romagna', 'atala', 'lodi',  'zaharieva', 'moritz', 'kraemer', 'pfaffenbach', 'terawan', 'agus', 'putranto', 'gualtieri', 'ramesh', 'raskar', 'amesh', 'adalja','corte', 'madera', 'baermarch', 'meselson', 'tuas','siouxsie', 'engen', 'tirol', 'salzburg', 'ciampino', 'fiumicino', 'oleksyk', 'nyayien', 'banang', 'golnaz', 'esfandiari', 'rfe/rl', 'kaszeta', 'kerckhoven', 'laatste', 'nieuws', 'hribar', 'balagh', 'naba']
foreign = ['...', '2,630', '^', 'ks', '—', '…', 'ecj', '–', 'von', 'der', 'leyen', 'draghi',  'ukhsa', '1-800-232-0233', '1-888-720-7489', 'kkmnow', '§', 'vaughan', 'ɔː', 'aɪ', 'vawn', 'yke', 'orf1ab','duomo', 'kannada', 'kazakh', 'kurmanji', 'kyrgyz', 'vue', '>', '+', 'kjzz', 'spot127', 'yediyurappa', 'mizoram', '#', '10.1038/s41586-023-06952-2', 'janna', 'jäe', 'tameka', 'kiribati', 'kosovo',  'tamil', 'nadu', 'telangana', 'medak', 'rangareddy',  '2,311', '04-5192519',  '•',  '0.12–0.16', '..', '1.9', '2.9',  'avolta', '05:08', '31.1', '22.9', '20.1', '19.6', '15.8', '13.2', '10.8', '10.5', '1,400', '8,700', 'nrevss', '229e', 'nl63', 'oc43', 'hku1', 'xbb', '44.2', '6:34:13', 'bnt162b2', '0.131', '0.043–0.218', 'siddaramaiah',  'megyn', '0:03', '1:55', 'vijayakanth', 'hmong', '102+', '116+', 'e18–e21', 'doi:10.2337/dc20-1872', 'önmez', '4.64', '±', '1.19', 'emd-15965', 'nsp3-4', 'ubl1-ubl2', 'ba.2', 'ba.4', 'ba.5','px200es', 'b.1.1.7', 'b.1.351', 'p.1', '15:09', 'ba.2.86', '$', '21.6', '£65', '2021–2022', '126.99', '175.99', 'siddaramaiah-led', 'ls-p2-vp8', '~15', '~2', '6,7', '892–903', 'tip3p', 'å', '×', '74.29', '5.71', 'xbb.1.9.2', '–51', '888-665-5865', '686-0802', 'fisr-2020ip-04249', '0.', '+853', '+90', 'ba.2·86', '61.8', '877-vax', '1,595,055', '296,626', '576,910', '627,773', '93,746', '11:01', 'gujarat', 'rajasthan', '17,379', '3,796','sanya', 'tedros', 'adhanom', 'ghebreyesus', '‘', '09:10', '08:05', 'file:2023-12-04', 'pmc9767341', 'ヴガールケラモフvs山本空良の', '試合に八百長発覚。許せない', 'anshu', 'covid-italy-rt-aa-200319_hpmain_16x9_992', 'ga4', '....', '2:53', '00:12', 'worldwide\u200b', '//odysee.com/', '16:21', '27th',  'फ़ोटो', 'और', 'वीडियो', ' weapon~1hrthusvrbu']

qcount=0
for cell1, cell2 in cell_obj:
    cellword1 = cell1.value
    for w in cellword1:
        if w=="\"":
            qcount += 1
            qcountt += 1
        
    cellword2 = cell2.value
    for w in cellword2:
        if w=="\"":
            qcount += 1
            qcountd += 1


i = 1

for cell1, cell2 in cell_obj:
    words = word_tokenize(cell1.value)
    
    c1 = sheet_obj.cell(row = i, column = 1)
    filtered= []
    #filtered_text = [t for t in words if not t.lower() in stopwords.words("english")]
    #c1.value = " ".join(filtered_text)
    for w in words:
        
        if w.lower() not in stopWords and w.lower() not in foreign and not w.isnumeric():
            if not w == "``" and not w=="''" and not w=='"': 
                for c in w:
                    
                    if not c.isalnum() and not c=="!" and not c=="?" and not c=='"' and not c==":":
                        w = w.replace(c, "")
                if w.isnumeric():
                    w = ""
                if not w == "":
                    filtered.append(w)
                    wordcount += 1
            else:
                w = "\""
                filtered.append(w)
                wordcount += 1
        elif w=="WHO":
            filtered.append(w)
            wordcount += 1
    c1.value = " ".join(filtered)
    
    filtered2= []
    words = word_tokenize(cell2.value)

    c2 = sheet_obj.cell(row = i, column = 2)   
    #filtered_text = [t for t in words if not t.lower() in stopwords.words("english")]
    #c2.value = " ".join(filtered_text)
    for w in words:
        
        if w.lower() not in stopWords and w.lower() not in foreign and not w.isnumeric():
            if not w == "``" and not w=="''" and not w=='"': 
                for c in w:
    
                    if not c.isalnum() and not c=="!" and not c=="?" and not c=='"' and not c==":":
                        w = w.replace(c, "")
                if w.isnumeric():
                    w = ""
                if not w == "":
                    filtered2.append(w)
                    wordcount += 1
            else:
                w = "\""
                filtered2.append(w)
                wordcount += 1
        elif w=="WHO":
            filtered2.append(w)
            wordcount += 1
            
    c2.value = " ".join(filtered2)
    i=i+1
    #print(cell1.value, cell2.value)
    

print("quotes: "+str(qcount))
print("quotes-title: "+str(qcountt))
print("quotes-desc: "+str(qcountd))
print("wordcount: "+str(wordcount))
#wb_obj.save("Database_cleaned.xlsx") 
#wb_obj.save("sample.xlsx") 
wb_obj.save("../../Data_Collection_03/02_Preprocessing/01_TitleDesc/02_TitleDesc_stop_punc_fw_removed_trash.xlsx") 