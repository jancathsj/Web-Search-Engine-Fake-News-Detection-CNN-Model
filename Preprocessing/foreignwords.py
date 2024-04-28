# -*- coding: utf-8 -*-
"""
Created on Wed May 31 20:37:03 2023

@author: Jan Catherine
"""

import openpyxl
from openpyxl import Workbook
import nltk
from nltk.tokenize import sent_tokenize, word_tokenize
#nltk.download('words')

wb_obj = openpyxl.load_workbook('../../Data_Collection_03/02_Preprocessing/01_TitleDesc/01_TitleDesc.xlsx')

sheet_obj = wb_obj.active 

row = sheet_obj.max_row
column = sheet_obj.max_column
rowcell = "B"+str(row)
cell_obj = sheet_obj['A1': rowcell]

englishwords = set(nltk.corpus.words.words())
foreign= []
i = 1

for cell1, cell2 in cell_obj:
    wordss = word_tokenize(cell1.value)
    
    c1 = sheet_obj.cell(row = i, column = 1)
   

    for w in wordss:
        
        if w.lower() not in englishwords:
            if w.lower() not in foreign and not w=="!" and not w=="?" and not w=="\"" and not w==":" and not w.isnumeric():
                foreign.append(w.lower())
            

    c2 = sheet_obj.cell(row = i, column = 2)   
    #filtered_text = [t for t in words if not t.lower() in stopwords.words("english")]
    #c2.value = " ".join(filtered_text)
    wordss = word_tokenize(cell2.value)
    for w in wordss:
        
        if w.lower() not in englishwords:
            if w.lower() not in foreign and not w=="!" and not w=="?" and not w=="\"" and not w==":" and not w.isnumeric():
                   foreign.append(w.lower())
    i=i+1
    #print(cell1.value, cell2.value)
    

print(foreign)

