#count capital words and punctuation marks

import openpyxl
from openpyxl import Workbook
import nltk
from nltk.tokenize import sent_tokenize, word_tokenize

# Give the location of the file 
# path = "gfg.xlsx"
  
# To open the workbook 
# workbook object is created 
wb_obj = openpyxl.load_workbook('../../Data_Collection_02/02_Preprocessing/01_TitleDesc/03_TitleDesc_cleaned_Real.xlsx') 
  
# Get workbook active sheet object 
# from the active attribute 
sheet_obj = wb_obj.active 

row = sheet_obj.max_row
column = sheet_obj.max_column
rowcell = "B"+str(row)
print(rowcell)

# Cell object is created by using 
# sheet object's cell() method. 
cell_obj = sheet_obj['A1': rowcell]

punc = ['!','?',':']
pc = [0,0,0]
pctitle = 0
pcdesc = 0
i = 1
capwords = []
capwordc = []
captitle = 0
capdesc = 0
capcmax = 0
capccur = 0
recorded = False
wordcount = 0

for cell1, cell2 in cell_obj:
    print(i)
    words = word_tokenize(cell1.value)
    c1 = sheet_obj.cell(row = i, column = 1)
    for w in words:
        wordcount += 1
        for c in w:
            mc = 0
            for m in punc:
                if c == punc[mc]:
                    pc[mc] = pc[mc]+1
                    pctitle = pctitle+1
                mc = mc + 1
        if w.isupper():
            capccur = 0
            recorded = False
            for cw in capwords:
                if w == cw:
                    capwordc[capccur] = capwordc[capccur] + 1
                    captitle = captitle + 1
                    recorded = True
                else:
                    capccur = capccur + 1
                    
            if not recorded:
                capwords.append(w)
                capwordc.append(0)
                capcmax += 1
                capwordc[capccur] = capwordc[capccur] + 1
                captitle = captitle + 1
                
    words = word_tokenize(cell2.value)
    c2 = sheet_obj.cell(row = i, column = 2)
    for w in words:
        wordcount += 1
        for c in w:
            mc = 0
            for m in punc:
                if c == punc[mc]:
                    pc[mc] = pc[mc]+1 
                    pcdesc = pcdesc+1 
                mc = mc + 1
        if w.isupper():
            capccur = 0
            recorded = False
            for cw in capwords:
                if w == cw:
                    capwordc[capccur] = capwordc[capccur] + 1
                    capdesc = capdesc + 1
                    recorded = True
                else:
                    capccur = capccur + 1
                    
            if not recorded:
                capwords.append(w)
                capwordc.append(0)
                capcmax += 1
                capwordc[capccur] = capwordc[capccur] + 1
                capdesc = capdesc + 1
    i=i+1


wb_new = Workbook()
# insert value in the cells
sheet_obj_new = wb_new.active 

rowcell_new = "B"+str(capcmax)

# Cell object is created by using 
# sheet object's cell() method. 
cell_obj_new = sheet_obj['A1': rowcell_new]
i2=1
temp = 0
for a in capwords:
 #   for cell1, cell2 in cell_obj_new:
     c1n = sheet_obj_new.cell(row = i2, column = 1)
     c2n = sheet_obj_new.cell(row = i2, column = 2)
     c1n.value = capwords[temp]
     c2n.value = capwordc[temp]
     temp += 1
     i2 += 1
print ("punc count:"+str(pc))
print ("punc count title: "+str(pctitle))
print ("punc count desc: "+str(pcdesc))
print ("capital count title: "+str(captitle))
print ("capital count desc: "+str(capdesc))
print ("wordcount: "+str(wordcount))

# save workbook as .xlsx file
#wb_new.save("Fake_uppercase.xlsx")
wb_new.save('../../Data_Collection_02/02_Preprocessing/01_TitleDesc/04_TitleDesc_Uppercase_Real.xlsx')



    